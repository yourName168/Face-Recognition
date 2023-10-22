from tkinter import messagebox
import cv2
import os
import numpy as np
from PIL import Image
import openpyxl


class FaceRegister:
    def __init__(self, data_path, recognizer_path, excel_path):
        self.recognizer = cv2.face_LBPHFaceRecognizer.create()
        self.data_path = data_path
        self.recognizer_path = recognizer_path
        self.excel_path = excel_path

    def collect_data(self):
        user_id = self.get_next_id()
        video = cv2.VideoCapture(0)

        facedetect = cv2.CascadeClassifier("haarcascade_frontalface_default.xml")
        count = 0

        while True:
            ret, frame = video.read()
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = facedetect.detectMultiScale(gray, 1.3, 5)

            for x, y, w, h in faces:
                count += 1
                cv2.imwrite(
                    os.path.join(self.data_path, f"User.{user_id}.{count}.jpg"),
                    gray[y : y + h, x : x + w],
                )
                cv2.rectangle(frame, (x, y), (x + w, y + h), (50, 50, 255), 1)

            cv2.imshow("Frame", frame)

            k = cv2.waitKey(1)
            if k == ord("q"):
                break
            if count > 200:
                break

        video.release()
        cv2.destroyAllWindows()
        messagebox.showinfo("Thông báo", "Thu thập dữ liệu thành công.")

    def train_model(self):
        imagePaths = [
            os.path.join(self.data_path, f) for f in os.listdir(self.data_path)
        ]
        faces = []
        IDs = []

        for imagePath in imagePaths:
            faceImg = Image.open(imagePath).convert("L")
            faceNp = np.array(faceImg, "uint8")
            user_id = int(os.path.split(imagePath)[-1].split(".")[1])

            faces.append(faceNp)
            IDs.append(user_id)

        if len(IDs) > 0:
            messagebox.showinfo("Thông báo", "Processing...")
            self.recognizer.train(faces, np.array(IDs))
            self.recognizer.save(self.recognizer_path)
            messagebox.showinfo("Thông báo", "Huấn luyện hoàn tất.")
        else:
            messagebox.showerror("Lỗi", "Không tìm thấy dữ liệu huấn luyện.")

    def save_data(self, name, student_id, dob, class_name):
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb.active
        next_id = self.get_next_id()
        ws.cell(row=next_id, column=1, value=next_id)
        ws.cell(row=next_id, column=2, value=name)
        ws.cell(row=next_id, column=3, value=student_id)
        ws.cell(row=next_id, column=4, value=dob)
        ws.cell(row=next_id, column=5, value=class_name)
        wb.save(self.excel_path)

    def get_next_id(self):
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb.active
        return ws.max_row + 1
