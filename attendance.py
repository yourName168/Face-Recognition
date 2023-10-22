import cv2
import openpyxl
import os
import shutil

class FaceRecognition:
    def __init__(self, face_detect_path, excel_path):
        self.video = cv2.VideoCapture(0)
        self.path = os.path.dirname(__file__)
        self.facedetect = cv2.CascadeClassifier(face_detect_path)
        self.excel_path = excel_path
        self.trainer_status = True
        self.recognizer = cv2.face_LBPHFaceRecognizer.create()
        trainer_file = os.path.join(self.path, "trainer.yml")
        try:
            self.recognizer.read(trainer_file)
        except Exception as e:
            print("Cảnh báo: Xảy ra lỗi khi đọc tệp 'trainer.yml'")
            self.trainer_status = False
            raise Exception("Không thể tạo đối tượng vì lỗi đọc tệp 'trainer.yml'")
        self.student_ids = self.read_ids_from_excel()

    def read_ids_from_excel(self):
        student_ids = {}
        try:
            workbook = openpyxl.load_workbook(os.path.join(self.path, "data1.xlsx"))
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) >= 3:
                    id, name, student_id = int(row[0]), row[1], row[2]
                    student_ids[id] = student_id
        except Exception as e:
            print(f"An error occurred while reading the Excel file: {e}")

        return student_ids

    def run(self):
        while True:

            ret, frame = self.video.read()
            if not ret:
                print("Không thể đọc khung hình từ camera.")
                break
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = self.facedetect.detectMultiScale(gray, 1.3, 5)
            for x, y, w, h in faces:
                serial, confident = self.recognizer.predict(gray[y : y + h, x : x + w])
                if confident < 50:
                    print(confident)
                    id = self.student_ids.get(serial)
                    cv2.putText(
                        frame,
                        f"{str(id).upper()}",
                        (x - 100, y + h - 300),
                        cv2.FONT_HERSHEY_COMPLEX,
                        1,
                        (0, 255, 0),
                        2,
                    )
                    cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)

                    self.update_status_in_excel(serial)

                else:
                    cv2.putText(
                        frame,
                        "Unknown",
                        (x, y + h - 300),
                        cv2.FONT_HERSHEY_COMPLEX,
                        1,
                        (0, 0, 255),
                        2,
                    )
                    cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 0, 255), 2)
                cv2.imshow("Face Recognition", frame)
            if cv2.waitKey(1) & 0xFF == ord("q"):
                break
        self.video.release()
        cv2.destroyAllWindows()

    def update_status_in_excel(self, student_id):
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == student_id:
                row_index = student_id
                if row[-1] == "V":
                    ws.cell(row=row_index, column=ws.max_column, value="X")
            wb.save(self.excel_path)
        self.create_backup()
        
    def create_backup(self):
        if os.path.exists(self.excel_path):
            backup_file = os.path.join(self.path, "backup_data1.xlsx")
            shutil.copyfile(self.excel_path, backup_file)
