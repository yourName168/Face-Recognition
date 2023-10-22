import tkinter as tk
import openpyxl
from tkinter import Toplevel, ttk
from datetime import datetime
from register import FaceRegister
from attendance import FaceRecognition
import threading
import os
import shutil


class FaceRecognitionApp:
    def __init__(self, root):
        self.root = root
        
        self.excel_path = os.path.join(os.path.dirname(__file__), "data1.xlsx")
        self.backup_excel_path = os.path.join(
            os.path.dirname(__file__), "backup_data1.xlsx"
        )
        self.create_backup()
        self.root.after(1000, self.update_table_periodically)
        self.face_register = FaceRegister(
            data_path=os.path.join(os.path.dirname(__file__), "data"),
            recognizer_path=os.path.join(os.path.dirname(__file__), "trainer.yml"),
            excel_path=self.excel_path,
        )
        self.root.title("Face Recognition")
        self.root.geometry("600x600")
        self.root.configure(bg="lightgrey")

        self.frame_label = tk.Label(
            self.root,
            text="Hệ thống Nhận diện Khuôn mặt",
            font=("Arial", 20, "bold"),
            bg="lightgrey",
            fg="black",
        )
        self.frame_label.pack(pady=20)

        self.register_button = tk.Button(
            self.root,
            text="Đăng ký",
            command=self.show_registration_dialog,
            font=("Arial", 14),
            bg="lightblue",
            fg="black",
            padx=20,
            pady=10,
        )
        self.register_button.pack(pady=10)

        self.attendance_button = tk.Button(
            self.root,
            text="Điểm danh",
            command=self.attendance,
            font=("Arial", 14),
            bg="lightgreen",
            fg="black",
            padx=20,
            pady=10,
        )
        self.attendance_button.pack(pady=10)

        self.attendance_process = None

        self.tree = ttk.Treeview(self.root)
        self.tree["columns"] = ("Tên", "Mã SV", "Ngày Sinh", "Status")
        self.tree.heading("#0", text="STT")
        self.tree.column("#0", anchor="center", width=50)
        self.tree.heading("Tên", text="Tên")
        self.tree.heading("Mã SV", text="Mã SV")
        self.tree.heading("Ngày Sinh", text="Ngày Sinh")
        self.tree.heading("Status", text="Status")

        self.tree.column("Tên", width=150)
        self.tree.column("Mã SV", width=100)
        self.tree.column("Ngày Sinh", width=150)
        self.tree.column("Status", width=100)

        self.tree.pack(expand=True, fill="both")

        self.tree.tag_configure("X", background="green")
        self.tree.tag_configure("V", background="red")

    def show_registration_dialog(self):
        registration_dialog = Toplevel(self.root)
        registration_dialog.title("Đăng ký")
        registration_dialog.geometry("300x400")
        registration_dialog.configure(bg="lightgrey")

        name_label = tk.Label(
            registration_dialog, text="Họ và tên:", font=("Arial", 12), bg="lightgrey"
        )
        name_label.pack(pady=10)
        name_entry = tk.Entry(registration_dialog, font=("Arial", 12))
        name_entry.pack(pady=5)

        student_id_label = tk.Label(
            registration_dialog,
            text="Mã sinh viên:",
            font=("Arial", 12),
            bg="lightgrey",
        )
        student_id_label.pack(pady=10)
        student_id_entry = tk.Entry(registration_dialog, font=("Arial", 12))
        student_id_entry.pack(pady=5)

        dob_label = tk.Label(
            registration_dialog,
            text="Ngày tháng năm sinh (dd/mm/yyyy):",
            font=("Arial", 12),
            bg="lightgrey",
        )
        dob_label.pack(pady=10)
        dob_entry = tk.Entry(registration_dialog, font=("Arial", 12))
        dob_entry.pack(pady=5)

        class_label = tk.Label(
            registration_dialog, text="Lớp:", font=("Arial", 12), bg="lightgrey"
        )
        class_label.pack(pady=10)
        class_entry = tk.Entry(registration_dialog, font=("Arial", 12))
        class_entry.pack(pady=5)

        submit_button = tk.Button(
            registration_dialog,
            text="Đăng ký",
            command=lambda: self.register_student(
                name_entry.get(),
                student_id_entry.get(),
                dob_entry.get(),
                class_entry.get(),
                registration_dialog,
            ),
            font=("Arial", 12),
            bg="lightblue",
            fg="black",
            padx=10,
            pady=5,
        )
        submit_button.pack(pady=10)

    def register_student(self, name, student_id, dob, class_name, registration_dialog):
        if name and student_id and dob and class_name:
            self.face_register.collect_data()
            self.face_register.train_model()
            self.face_register.save_data(name, student_id, dob, class_name)
            now = datetime.now()
            column_name = now.strftime("%Y-%m-%d %H")
            self.create_status_column(column_name)
            registration_dialog.destroy()
            self.create_backup()
            self.show_table()
    def update_status(self,column_name):
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb.active        
        status_column = ws.max_column+1
        ws.cell(row=1, column=ws.max_column + 1, value=column_name)        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=status_column, max_col=status_column):
            for cell in row:
                if cell.value != "X":
                    cell.value = "V"
        wb.save(self.excel_path)

    def create_status_column(self, column_name):
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb.active
        column_exists = False
        for cell in ws[1]:
            if cell.value == column_name:
                column_exists = True
                break
        if not column_exists:
            self.update_status(column_name)

    def attendance(self):
        if (
            self.attendance_process is None
            or self.attendance_process.poll() is not None
        ):
            now = datetime.now()
            column_name = now.strftime("%Y-%m-%d %H")
            self.update_table_periodically()
            try:
                self.face_attendance = FaceRecognition(
                    os.path.join(
                        os.path.dirname(__file__), "haarcascade_frontalface_default.xml"
                    ),
                    self.excel_path,
                )
            except:
                self.show_registration_required_message()
                return
            self.create_status_column(column_name)
            self.face_attendance.video.release()
            self.face_attendance = FaceRecognition(
                os.path.join(
                    os.path.dirname(__file__), "haarcascade_frontalface_default.xml"
                ),
                self.excel_path,
            )
            attendance_thread = threading.Thread(target=self.face_attendance.run)
            attendance_thread.start()

    def show_registration_required_message(self):
        registration_message = tk.Label(
            self.root,
            text="Không có dữ liệu hiển thị. Vui lòng đăng kí trước khi điểm danh",
            font=("Arial", 12),
            bg="lightgrey",
        )
        registration_message.pack(pady=10)
        self.root.after(5000, registration_message.destroy)

    def update_table_periodically(self):
        self.show_table()
        self.root.after(1000, self.update_table_periodically)
        return False

    def show_table(self):
        try:
            wb = openpyxl.load_workbook(self.backup_excel_path)
        except:
            wb = openpyxl.load_workbook(self.excel_path)
        ws = wb.active
        row_number = 1
        self.tree.delete(*self.tree.get_children())

        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) >= 5:
                name, maSV, date, status = row[1], row[2], row[3], row[-1]

                if status == "X":
                    tags = ("X",)
                else:
                    tags = ("V",)

                self.tree.insert(
                    "",
                    "end",
                    text=row_number,
                    values=(name, maSV, date, status),
                    tags=tags,
                )
                row_number += 1

        wb.close()

    def create_backup(self):
        if os.path.exists(self.excel_path):
            backup_file = os.path.join(os.path.dirname(__file__), "backup_data1.xlsx")

            shutil.copyfile(self.excel_path, backup_file)


if __name__ == "__main__":
    root = tk.Tk()
    app = FaceRecognitionApp(root)
    root.mainloop()
